# -*- coding: utf-8 -*-
# pip install click --upgrade
import librosa
from flask import Flask, request, jsonify, make_response
import os
import alitts
from pydub import AudioSegment
import requests
import time
import json
import configparser
from difflib import SequenceMatcher
import http.client
from codecs import encode
from threading import Timer
import uuid
import subprocess
import pandas as pd
import openpyxl
import difflib
import asyncio
import websockets
from pythonosc import udp_client
import re
import pypinyin
from pypinyin import Style
import datetime
from fuzzywuzzy import fuzz
from sparkai.llm.llm import ChatSparkLLM, ChunkPrintHandler
from sparkai.core.messages import ChatMessage
import time
from flask import request
import queue
import threading

config = configparser.ConfigParser()
config.read('./secrets.ini')

#星火认知大模型Spark Max的URL值，其他版本大模型URL值请前往文档（https://www.xfyun.cn/doc/spark/Web.html）查看
SPARKAI_URL = config.get('xfzhinengti', 'SPARKAI_URL')
#星火认知大模型调用秘钥信息，请前往讯飞开放平台控制台（https://console.xfyun.cn/services/bm35）查看
SPARKAI_APP_ID = config.get('xfzhinengti', 'SPARKAI_APP_ID')
SPARKAI_API_SECRET = config.get('xfzhinengti', 'SPARKAI_API_SECRET')
SPARKAI_API_KEY = config.get('xfzhinengti', 'SPARKAI_API_KEY')
#星火认知大模型Spark Max的domain值，其他版本大模型domain值请前往文档（https://www.xfyun.cn/doc/spark/Web.html）查看
SPARKAI_DOMAIN = 'generalv3.5'
spark = ChatSparkLLM(
        spark_api_url=SPARKAI_URL,
        spark_app_id=SPARKAI_APP_ID,
        spark_api_key=SPARKAI_API_KEY,
        spark_api_secret=SPARKAI_API_SECRET,
        spark_llm_domain=SPARKAI_DOMAIN,
        streaming=True,
    )

doubao_apikey=config.get('doubao', 'apikey')
doubao_botname=config.get('doubao', 'botmodel')

# 创建一个队列
data_queue = queue.Queue()
# 处理函数
def process_queue():
    while True:
        # 从队列中获取数据
        data = data_queue.get()
        if data is None:  # 用于结束线程
            break
        # 处理数据
        print(f"Processing: {data}")
        # 开始说话
        total_length=a2fspeakout(data)
        time.sleep(total_length)
        timeout = float(total_length) 
        start_timer(key, timeout)   
        data_queue.task_done()
# 启动后台线程
thread = threading.Thread(target=process_queue, daemon=True)
thread.start()

def get_fuzzy_pinyin(text):
    """
    获取中文字符串text的模糊拼音码
    """
    # 将中文字符串转换为拼音列表
    pinyin_list = pypinyin.lazy_pinyin(text, style=Style.NORMAL)
    # 将拼音列表连接成字符串
    fuzzy_pinyin = ''.join(pinyin for py in pinyin_list for pinyin in py)
    return fuzzy_pinyin

def match_fuzzy_pinyin(text, pinyin_str):
    """
    判断中文字符串text的拼音是否与pinyin_str模糊匹配
    "xiaonanxiaonan", "xiaolanxiaolan", "xiaonanxiaolan", "xiaolanxiaonan"
    """
    # 将中文字符串转换为拼音列表
    pinyin_list = pypinyin.lazy_pinyin(text)
    # 构造正则表达式模式
    patterns = [
        '.*'.join(['', 'shu', 'zhi', 'yuan', 'gong', '']),
        '.*'.join(['', 'su', 'zhi', 'yuan', 'gong', '']),
        '.*'.join(['', 'shu', 'zi', 'yuan', 'gong', '']),
        '.*'.join(['', 'su', 'zi', 'yuan', 'gong', ''])
    ]
    # 判断拼音字符串是否模糊匹配
    return any(re.match(pattern, pinyin_str, re.IGNORECASE) for pattern in patterns)

app = Flask(__name__)
app.secret_key = 'daowifhsefighsaofhia' # 这里不用改
wav_name = "test.wav"
usd_file_name = "DefaultOfficialInstance.usd"
usd_absolute_path = os.path.abspath(usd_file_name)
a2fserverurl='http://127.0.0.1:10246'

answer_sentence="" # 判断文本相似度
request_end_time = {} # 判断打断
API_KEY = config.get('wenxin', 'apikey') # 文心一言密钥
SECRET_KEY = config.get('wenxin', 'appsecret') # 文心一言密钥

timers = {}
key = uuid.uuid4()
# 全局变量，用于存储句子的长度
MIN_CHINESE_CHAR_COUNT = 20
accumulated_message = ""

async def send_message_to_websocket(message):
    async with websockets.connect('ws://localhost:60001') as websocket:
        await websocket.send(message)  # 发送消息给服务器
        print("Message sent")

def send_message(message):
    asyncio.run(send_message_to_websocket(message))

def timer_expired(key):
    print(f"说话结束")

def start_timer(key, timeout):
    timers[key] = Timer(timeout, timer_expired, args=(key,))
    timers[key].start()

def delay_response(delay, response):
    """延时返回响应"""
    time.sleep(delay)
    return make_response(response)

def a2finit():
    global usd_absolute_path
    global wav_name
    print("a2f嘴型usd文件位置："+usd_absolute_path)
    url = a2fserverurl+'/A2F/USD/Load'
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {
        'file_name': usd_absolute_path
    }
    response = requests.post(url, headers=headers, json=data)
    wav_absolute_pathdir = os.path.abspath(wav_name).replace("\\"+wav_name,"")
    url = a2fserverurl+'/A2F/Player/SetRootPath'
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {
        'a2f_player': '/World/audio2face/Player',
        'dir_path': wav_absolute_pathdir
    }
    response = requests.post(url, headers=headers, json=data)


def a2fpause():
    total_length=a2fspeakout(",")
    time.sleep(1)
    my_dict = {
        "type": "bizui",
        "content": "你好我在"
    }
    json_string = json.dumps(my_dict, ensure_ascii=False)
    result = json_string
    send_message(result)
    return ""

def a2fspeakout(output):
    global answer_sentence
    global wav_name
    wav_file=wav_name
    answer_sentence=output
    alitts.speakword(wav_file,output)
    #huansheng(wav_file,output)
    # 计算音频总长度，秒
    total_length = get_duration(wav_file)
    audio = AudioSegment.from_file(wav_file)
    length = len(audio) / 1000 # 获取的长度单位是毫秒，转换为秒钟
    url = a2fserverurl+'/A2F/Player/SetTrack'
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    dat2a = {
        'a2f_player': '/World/audio2face/Player',
        'file_name': wav_name,
        'time_range': [0, -1]
    }
    response = requests.post(url, headers=headers, json=dat2a)
    url = a2fserverurl+'/A2F/Player/Play'
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    dat2a = {
        'a2f_player': '/World/audio2face/Player'
    }
    response = requests.post(url, headers=headers, json=dat2a)
    return total_length

# 问答
# 请介绍下我们公司？ 我们公司是与云计算伴生的一项基于超级计算机系统对外提供计算资源,存储资源等服务的机构或单位，以高性能计算机为基础面向各界提供高性能计算服务。
# 我还想了解更多？我们公司致力于为各行各业提供高性能计算服务，利用高性能计算机系统提供计算资源,存储资源等解决方案。我们的目标是通过云计算技术帮助客户实现更快,更强大的计算能力，以推动科学研究,工程设计和商业创新的发展。我们的团队拥有丰富的经验和专业知识，致力于为客户提供可靠,安全,高效的计算服务，以满足不断增长的需求。

# 获取文心一言token
def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": API_KEY, "client_secret": SECRET_KEY}
    return str(requests.post(url, params=params).json().get("access_token"))

# 获取wav时长
def get_duration(file_path):
    try:
        y, sr = librosa.load(file_path, sr=None) # sr=None 保持原始采样率
        duration = librosa.get_duration(y=y, sr=sr)
        return duration
    except Exception as e:
        print(f"Error with librosa: {e}")
        return None

# 清理状态初始化
@app.route('/clearstatus', methods=['POST'])
def clearstatus():
    global answer_sentence
    answer_sentence="" # 判断文本相似度
    data = request.form
    # 接收 JSON 数据
    # F:/audio2face-2023.1.1/exts/omni.audio2face.player_deps/deps/audio2face-data/tracks/
    # {"message":"你好吗"}
    data_message = data.get("message")
    # 根据中文TTS生成wav文件
    output = data_message
    # 开始说话
    total_length=a2fspeakout(output)
    data={}
    data["message"]=str(data_message)
    data["record_time"]=total_length
    return jsonify(data)

# 模拟语音说话
@app.route('/test', methods=['POST'])
def test():
    data = request.form
    data_message = data.get("message")
    # 根据中文TTS生成wav文件
    output = data_message
    # 开始说话
    total_length=a2fspeakout(output)
    data = {}
    data["message"]=data_message
    data["record_time"]=total_length
    return jsonify(data)

# 直接说话
@app.route('/apppost', methods=['POST'])
def speak():
    global answer_sentence
    global request_end_time
    global timers
    global key
    data = request.form
    data_message = data.get("message")
    
    if key in timers and timers[key].is_alive():
        if "小智小智" in str(data_message):
            data_message="按钮打断闭嘴"
            print("按钮打断闭嘴")
            timers = {}
            a2fpause()
            data = {}
            data["message"]=data_message
            data["record_time"]=0
            return jsonify(data)
        else:
            print("正在说话，开始打断")
            timers = {}
            output = "你好我在"
            # 开始说话
            total_length=a2fspeakout(output)
            data = {}
            data["message"]=data_message
            data["record_time"]=total_length
            return jsonify(data)
            
    else:
        if "小智小智" in str(data_message):
            data_message="按钮打断闭嘴"
            print("按钮打断闭嘴")
            timers = {}
            a2fpause()
            data = {}
            data["message"]=data_message
            data["record_time"]=0
            return jsonify(data)
        else:
            # 根据中文TTS生成wav文件
            output = data_message
            # 开始说话
            total_length=a2fspeakout(output)
            data = {}
            data["message"]=data_message
            data["record_time"]=total_length
            return jsonify(data)

# 文心一言回答 + 知识库回答
@app.route('/wenxin', methods=['POST'])
# @timer_limit
def wenxin():
    global answer_sentence
    global timers
    global key
    data = request.form
    data_message = data.get("message")
    print(data_message)
    # 判断文本相似度
    similarity_ratio = SequenceMatcher(None, answer_sentence, data_message).ratio()
    if similarity_ratio >= 0.6:
        return jsonify({'error': '答案回声传入，并非问题'}), 400
    else:
        if "天气" in str(data_message):
            print("提问文心一言3.5")
            url = "https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/completions?access_token=" + get_access_token()
            payload = json.dumps({
                "messages": [
                    {
                        "role": "user",
                        "content": "用最简洁的语言回答“"+data_message+"”这个问题"
                    }
                ],
                "temperature": 0.95,
                "top_p": 0.8,
                "penalty_score": 1,
                "disable_search": False,
                "enable_citation": False,
                "response_format": "text"
            })
            headers = {
                'Content-Type': 'application/json'
            }
            res = requests.request("POST", url, headers=headers, data=payload).json()
            output = res['result']
            output = output.replace("℃","摄氏度")
            output = output.replace("\\\\","到")
            output = output.replace("~","到")
            output = output.replace("-","到")
            output = output.replace("到到","到")
            output = output.replace("**","")
            output = output.replace("\\","")
            videoimg=""

        elif "日期" in str(data_message) and "今" in str(data_message):
            print("提问文心一言3.5")
            today = datetime.date.today()
            output = "今天日期是{}年{}月{}日".format(today.year, today.month, today.day)
            videoimg=""

        elif "几日几号" in str(data_message) and "今" in str(data_message):
            print("提问文心一言3.5")
            today = datetime.date.today()
            output = "今天日期是{}年{}月{}日".format(today.year, today.month, today.day)
            videoimg=""
        
        elif "几号" in str(data_message) and "今" in str(data_message):
            print("提问文心一言3.5")
            today = datetime.date.today()
            output = "今天日期是{}年{}月{}日".format(today.year, today.month, today.day)
            videoimg=""
                    
        elif "星期几" in str(data_message) and "今" in str(data_message):
            print("提问文心一言3.5")
            weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
            today = datetime.date.today()
            weekday = weekdays[today.weekday()]
            output = "今天是{}年{}月{}日,{}".format(today.year, today.month, today.day,weekday)
            videoimg=""

        elif "礼拜几" in str(data_message) and "今" in str(data_message):
            print("提问文心一言3.5")
            weekdays = ["礼拜一", "礼拜二", "礼拜三", "礼拜四", "礼拜五", "礼拜六", "礼拜天"]
            today = datetime.date.today()
            weekday = weekdays[today.weekday()]
            output = "今天是{}年{}月{}日,{}".format(today.year, today.month, today.day,weekday)
            videoimg=""
        else:
            # 提示用户输入问题
            user_question = data_message
            # 本地知识库模糊搜索
            a = allzskdata.question.apply(lambda user: fuzz.ratio(user, user_question))
            a = a.nlargest(5).reset_index()
            a.columns = ["question", "similar"]
            a.question = allzskdata.question[a.question].values
            df = pd.DataFrame(a)
            first_row_values = df.iloc[0, [0, 1]]
            print(float(first_row_values['similar']))
            if float(first_row_values['similar'])>40.0:
                df = pd.DataFrame(allzskdata)
                result = df.loc[df['question'] == first_row_values['question'], 'answer'].values[0]
                output=result
                videoimg = df.loc[df['question'] == first_row_values['question'], 'videoimg'].values[0]
                videoimg=videoimg

            else:
                print("提问文心一言")
                url = "https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/eb-instant?access_token=" + get_access_token()
                s=data_message
                # 注意message必须是奇数条
                payload = json.dumps({
                "messages": [
                    {
                        "role": "user",
                        "content": "用最简洁的语言回答“"+s+"”这个问题"
                    }
                ]
                })
                headers = {
                    'Content-Type': 'application/json'
                }
                res = requests.request("POST", url, headers=headers, data=payload).json()
                # 根据中文TTS生成wav文件
                print(res)
                output = res['result']
                output = output.replace("℃","摄氏度")
                output = output.replace("\\\\","到")
                output = output.replace("~","到")
                output = output.replace("-","到")
                output = output.replace("到到","到")
                output = output.replace("**","")
                output = output.replace("\\","")
                videoimg=""

        # 开始说话
        total_length=a2fspeakout(output)
        timeout = float(total_length) 
        start_timer(key, timeout)            
        data = {}
        data["message"]=output
        data["record_time"]=total_length
        if len(videoimg) > 0:
            data["videoimg"]="UpdateDataFromUe4('"+videoimg+"')"
        else:
            data["videoimg"]="UpdateDataFromUe4('"+"stop"+"')"
        return jsonify(data)

# doubao回答
@app.route('/doubao', methods=['POST'])
# @timer_limit
def doubao():
    global answer_sentence
    global timers
    global key
    data = request.form
    data_message = data.get("message")
    # 提示用户输入问题
    user_question = data_message
    print("提问豆包大模型")
    url = 'https://ark.cn-beijing.volces.com/api/v3/bots/chat/completions'
    headers = {
        'Authorization': 'Bearer '+doubao_apikey,
        'Content-Type': 'application/json'
    }
    data = {
        "model": doubao_botname,
        "stream": False,
        "stream_options": {"include_usage": True},
        "messages": [
            {
                "role": "user",
                "content": user_question
            }
        ]
    }
    response = requests.post(url, headers=headers, json=data)
    response_data = response.json()
    content = response_data['choices'][0]['message']['content']
    print(content)
    output=content
    # 开始说话
    total_length=a2fspeakout(output)
    timeout = float(total_length) 
    start_timer(key, timeout)            
    data = {}
    data["message"]=output
    data["record_time"]=total_length
    return jsonify(data)


# 讯飞智能体回答
@app.route('/xf_zhinengti', methods=['POST'])
# @timer_limit
def xf_zhinengti():
    global answer_sentence
    global timers
    global key
    global accumulated_message
    data = request.form
    data_message = data.get("message")
    # 提示用户输入问题
    user_question = data_message
    print("提问讯飞智能体")
    data = request.form
    data_message = data.get("message")
    messages = [ChatMessage(
        role="user",
        content=data_message
    )]
    a = spark.stream(messages)
    for message in a:
        print(message)
        # 累积消息
        if "additional_kwargs" not in str(message) :
            message=str(message).replace("\\n", "").replace("'", "").replace("content=", "")
            accumulated_message += str(message) 
            # 检查累积消息是否多于20个中文字符
            if len(accumulated_message) >= MIN_CHINESE_CHAR_COUNT:
                data_queue.put(accumulated_message)
                accumulated_message = ""  # 清空累积消息
    # url = 'https://ark.cn-beijing.volces.com/api/v3/bots/chat/completions'
    # headers = {
    #     'Authorization': 'Bearer df597f4e-3f9d-4cfe-968f-0bade2f7b79a',
    #     'Content-Type': 'application/json'
    # }
    # data = {
    #     "model": "bot-20240921105732-frfdn",
    #     "stream": False,
    #     "stream_options": {"include_usage": True},
    #     "messages": [
    #         {
    #             "role": "user",
    #             "content": user_question
    #         }
    #     ]
    # }
    # response = requests.post(url, headers=headers, json=data)
    # response_data = response.json()
    # content = response_data['choices'][0]['message']['content']
    # print(content)
    # output=content
    # # 开始说话
    # total_length=a2fspeakout(output)
    # timeout = float(total_length) 
    # start_timer(key, timeout)            
    # data = {}
    # data["message"]=output
    # data["record_time"]=total_length
    return jsonify(data)



# 星火大模型回答
@app.route('/xinghuo', methods=['POST'])
def xinghuo():
    data = request.form
    data_message = data.get("message")
    print("提问星火3.5")
    content="用最简洁的语言回答“"+data_message+"”这个问题"
    command = ['python', './xhbingmodel35.py', content]
    output = subprocess.check_output(command, universal_newlines=True)
    # 开始说话
    total_length=a2fspeakout(output)
    timeout = float(total_length) 
    start_timer(key, timeout)
    data = {}
    data["message"]=output
    data["record_time"]=total_length
    return jsonify(data)


@app.route('/getexcel', methods=['POST'])
def getexcel():
    data = request.form
    data_message = data.get("message")
    data = {}
    concatenated_string = ";".join(questions)
    data["content"]=concatenated_string
    data["count_no"]=excel_length
    print(data)
    return jsonify(data)

def huansheng(wav_file,message):
    speakername="Yennefer"
    output=message
    url = "http://127.0.0.1:5000/voice?model_id=0&speaker_name="+speakername+"&sdp_ratio=0.2&noise=0.2&noisew=0.9&length=1&language=ZH&auto_translate=false&auto_split=false&emotion=&style_weight=0.7"
    payload={"text":output}
    files=[
    ]
    headers = {
    'User-Agent': 'Apifox/1.0.0 (https://apifox.com)'
    }
    response = requests.request("POST", url, headers=headers, data=payload, files=files)
    # 将文件流写入到 WAV 文件
    with open(wav_file, "wb") as f:
        f.write(response.content)
    print("File saved as output.wav")

if __name__ == '__main__':
    print("加载本地知识库")
    # 加载Excel文件
    workbook = openpyxl.load_workbook('知识库问答模板-骏宇文博-4ab96e6a-4749-4495-adbb-3c217441e161.xlsx')
    sheet = workbook.active
    # 初始化问题和答案列表
    questions = []
    answers = []
    videoimg_list = []
    excel_length = 0
    # 遍历Excel中的数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], str) and isinstance(row[1], str):
            questions.append(row[0].replace("\n", ""))
            answers.append(row[1])
            videoimg_list.append("stop")
            excel_length=excel_length+1
    print("加载本地图文知识库")
    workbook = openpyxl.load_workbook('数字人图文知识库.xlsx')
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], str) and isinstance(row[1], str) and isinstance(row[2], str):
            questions.append(row[0].replace("\n", ""))
            answers.append(row[1])
            videoimg_list.append(row[2])
            excel_length=excel_length+1

    # 构建本地知识库
    allzskdata = pd.DataFrame({'question': questions, 'answer': answers, 'videoimg':videoimg_list})
    a2finit()
    # 开始说话
    total_length=a2fspeakout("初始化完成")
    #app.json.ensure_ascii = False # 解决中文json乱码
    app.run(host='0.0.0.0',port=5000, debug=True)

